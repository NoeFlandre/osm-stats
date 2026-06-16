"""Persist per-pipeline base key families as CSV and XLSX.

This is the shared writer used by both the filter-first and
standardize-first paths. The output schema and the XLSX formatting
are identical across paths; only the source medoid CSV (and the
output path) differ.

The XLSX layout:

* column 1: ``keep`` (the user fills in their verdict: ``keep``,
  ``drop``, ``?``, free text). Starts empty. This is the column the
  maintainer will use to manually classify each base key for the
  env/agri study.
* column 2: ``base_key``
* column 3: ``cluster_count``
* column 4: ``total_count_all``
* column 5: ``representative_medoids`` (the top-N medoids joined)

The first column is frozen so the user's verdict stays visible while
they scroll right, and the auto-filter is enabled on the full data
range so they can filter by verdict or by cluster_count.
"""
from __future__ import annotations

from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Schema the user sees in the XLSX. The first column is the verdict
# they fill in; the rest is the data they judge against.
XLSX_COLUMNS = [
    "keep",                 # A: user verdict
    "base_key",             # B
    "cluster_count",        # C
    "total_count_all",      # D
    "representative_medoids",  # E
]

XLSX_COLUMN_WIDTHS = {
    "A": 14,   # keep
    "B": 28,   # base_key
    "C": 14,   # cluster_count
    "D": 18,   # total_count_all
    "E": 90,   # representative_medoids
}

CSV_COLUMNS = [
    "base_key",
    "cluster_count",
    "total_count_all",
    "representative_medoids",
]


def save_base_key_families_csv(
    profile: pd.DataFrame, output_path: Union[str, Path]
) -> Path:
    """Write the per-base-key family table as CSV (no ``keep`` column).

    Sorted by ``total_count_all`` descending to match the cluster
    profile's ordering. Returns the path that was written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out = profile[CSV_COLUMNS].copy()
    out.to_csv(output_path, index=False)
    return output_path


def save_base_key_families_xlsx(
    profile: pd.DataFrame,
    output_path: Union[str, Path],
    sheet_name: str,
) -> Path:
    """Write the per-base-key family table as XLSX with a ``keep``
    classification column at the front.

    See module docstring for the layout. Returns the path that was
    written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = profile[CSV_COLUMNS].copy()
    df["keep"] = ""
    df = df[XLSX_COLUMNS]
    df.to_excel(output_path, index=False, sheet_name=sheet_name)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    # Freeze the first column AND the header row so the verdict
    # always stays visible while the user scrolls right.
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.row_dimensions[1].height = 22

    for col_letter, w in XLSX_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    wb.save(output_path)
    return output_path
