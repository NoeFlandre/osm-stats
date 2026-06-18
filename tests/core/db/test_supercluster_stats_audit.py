"""Audit tests for the supercluster element-type stats on real data.

These tests independently re-compute the per-supercluster stats from
the raw input files (cluster_memberships.csv, cluster_medoids.csv,
and taginfo.sqlite) and compare against the function's output. The
goal is to verify that the headline numbers are 100 % correct by
computing them two different ways and asserting they agree.

The tests use the real paths on the local Seagate M3 volume, which
holds the source DB and the standardize-first cache. They are
gated by the ``has_real_data`` fixture so they skip cleanly when
the data is unavailable. A session-scoped fixture runs the
expensive 170s function once and shares the result across all
tests in this module.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
import pytest

from src.core.db.element_type_stats import POLYGON_FRIENDLY_THRESHOLD
from src.core.db.supercluster_element_type_stats import supercluster_element_type_stats


SEAGATE = Path("/Volumes/Seagate M3")
SOURCE_DB = SEAGATE / "taginfo.sqlite"
SOURCE_CACHE = SEAGATE / "tag_features_standardize_first.sqlite"
MEMBERSHIPS = Path("output/standardize_first/tfidf/cluster_memberships.csv")
MEDOIDS = Path("output/standardize_first/tfidf/cluster_medoids.csv")
XLSX = Path("output/standardize_first/tfidf/base_key_families.xlsx")


def _has_real_data() -> bool:
    return all(
        p.exists() for p in (SOURCE_DB, SOURCE_CACHE, MEMBERSHIPS, MEDOIDS, XLSX)
    )


pytestmark = pytest.mark.skipif(
    not _has_real_data(),
    reason="real source DB / cache / outputs not available",
)


# --- shared session-scoped fixtures ----------------------------------------

@pytest.fixture(scope="module")
def yes_labels() -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX); ws = wb.active
    out: list[str] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        keep, bk = r[0], r[1]
        if bk is None:
            continue
        if str(keep or "").strip().lower() == "yes":
            out.append(str(bk))
    return out


@pytest.fixture(scope="module")
def cluster_to_base_key() -> dict[int, str]:
    from src.core.features.base_key import parse_base_key
    df = pd.read_csv(MEDOIDS)
    return {
        int(r.cluster_id): parse_base_key(r.medoid_feature)
        for r in df.itertuples()
    }


@pytest.fixture(scope="module")
def source_db_aggregate() -> pd.DataFrame:
    """One row per (LOWER(TRIM(key)), LOWER(TRIM(value))) with the
    per-element-type columns. Uses the cache as a JOIN filter to avoid
    the 300s+ full scan of the 192M-row source DB. The TRIM matches
    the cache's STANDARDIZED_KEY_EXPR / STANDARDIZED_VALUE_EXPR
    (see src/core/storage/cache.py).
    """
    con = sqlite3.connect(SOURCE_DB)
    try:
        con.execute(f"ATTACH DATABASE '{SOURCE_CACHE}' AS cache")
        df = pd.read_sql_query(
            """
            SELECT
                LOWER(TRIM(t.key))    AS key,
                LOWER(TRIM(t.value))  AS value,
                SUM(t.count_nodes)     AS count_nodes,
                SUM(t.count_ways)      AS count_ways,
                SUM(t.count_relations) AS count_relations
            FROM main.tags AS t
            JOIN cache.tag_features AS c
              ON c.key = LOWER(TRIM(t.key))
             AND c.value = LOWER(TRIM(t.value))
            GROUP BY LOWER(TRIM(t.key)), LOWER(TRIM(t.value))
            """,
            con,
        )
        con.execute("DETACH DATABASE cache")
    finally:
        con.close()
    df["count_all"] = (
        df["count_nodes"].fillna(0)
        + df["count_ways"].fillna(0)
        + df["count_relations"].fillna(0)
    )
    return df


@pytest.fixture(scope="module")
def supercluster_df(yes_labels, cluster_to_base_key) -> pd.DataFrame:
    """Run the function once. ~170s on the full 192M-row source DB."""
    return supercluster_element_type_stats(
        yes_labels, MEDOIDS, MEMBERSHIPS, SOURCE_DB, source_cache=SOURCE_CACHE,
    )


@pytest.fixture(scope="module")
def real_memberships(yes_labels, cluster_to_base_key) -> pd.DataFrame:
    """Cluster memberships restricted to real clusters whose medoid's
    base key is in the yes set.
    """
    yes_set = set(yes_labels)
    mem = pd.read_csv(MEMBERSHIPS)
    real = mem[mem["cluster_id"] != -1].copy()
    real["cluster_id"] = real["cluster_id"].astype("int64")
    real["supercluster_bk"] = real["cluster_id"].map(cluster_to_base_key)
    real = real.dropna(subset=["supercluster_bk"])
    return real[real["supercluster_bk"].isin(yes_set)]


# --- tests -----------------------------------------------------------------

def test_xlsx_has_exactly_157_yes_labels(yes_labels) -> None:
    """The user labeled 157 base keys as 'yes' (broader than the prior
    44 env/agri subset). Lock the count down."""
    assert len(yes_labels) == 157, f"expected 157 yes labels, got {len(yes_labels)}"


def test_count_all_per_supercluster_matches_independent_memberships_sum(
    supercluster_df, real_memberships, yes_labels,
) -> None:
    """For each supercluster, the function's count_all equals the sum
    of count_all in cluster_memberships for clusters whose medoid
    has that base key. Computed two independent ways."""
    actual = supercluster_df.set_index("base_key")["count_all"].astype("int64")
    expected = real_memberships.groupby("supercluster_bk")["count_all"].sum().astype("int64")

    assert set(actual.index) == set(expected.index), (
        f"keys differ: only-in-fn={set(actual.index)-set(expected.index)}, "
        f"only-in-sum={set(expected.index)-set(actual.index)}"
    )
    diffs = (actual - expected).abs()
    assert (diffs == 0).all(), (
        f"count_all mismatches in {(diffs>0).sum()} superclusters: "
        f"{diffs[diffs>0].head().to_dict()}"
    )


def test_n_clusters_per_supercluster_matches_independent_count(
    supercluster_df, real_memberships,
) -> None:
    """n_clusters = count of distinct cluster_ids per supercluster."""
    actual = supercluster_df.set_index("base_key")["n_clusters"].astype("int64")
    expected = real_memberships.groupby("supercluster_bk")["cluster_id"].nunique().astype("int64")

    assert set(actual.index) == set(expected.index)
    diffs = (actual - expected).abs()
    assert (diffs == 0).all(), (
        f"n_clusters mismatches: {diffs[diffs>0].head().to_dict()}"
    )


def test_n_tags_per_supercluster_matches_independent_count(
    supercluster_df, real_memberships,
) -> None:
    """n_tags = count of member rows per supercluster."""
    actual = supercluster_df.set_index("base_key")["n_tags"].astype("int64")
    expected = real_memberships.groupby("supercluster_bk").size().astype("int64")

    assert set(actual.index) == set(expected.index)
    diffs = (actual - expected).abs()
    assert (diffs == 0).all(), (
        f"n_tags mismatches: {diffs[diffs>0].head().to_dict()}"
    )


def test_element_type_split_per_supercluster_matches_independent_sql(
    supercluster_df, real_memberships, source_db_aggregate,
) -> None:
    """For each supercluster, count_nodes/ways/relations equals the
    SUM of source-DB element-type counts for its cluster members.
    """
    members = real_memberships[["supercluster_bk", "key", "value"]].copy()
    # Note: do NOT lowercase -- the cluster_memberships and the
    # source-DB aggregate are both SQLite LOWER+TRIM, and SQLite's
    # LOWER is ASCII-only. Re-lowercasing in Python with str.lower()
    # would break non-ASCII matches (e.g. Cyrillic, German umlauts).
    members = members.drop_duplicates()

    joined = members.merge(
        source_db_aggregate[["key", "value", "count_nodes", "count_ways", "count_relations"]],
        on=["key", "value"],
        how="left",
    )
    for c in ("count_nodes", "count_ways", "count_relations"):
        joined[c] = joined[c].fillna(0).astype("int64")
    expected = (
        joined.groupby("supercluster_bk", as_index=False)[["count_nodes", "count_ways", "count_relations"]].sum()
    )

    for col in ("count_nodes", "count_ways", "count_relations"):
        actual = supercluster_df.set_index("base_key")[col].astype("int64")
        exp = expected.set_index("supercluster_bk")[col].astype("int64")
        assert set(actual.index) == set(exp.index), f"{col}: index differs"
        diffs = (actual - exp).abs()
        assert (diffs == 0).all(), f"{col} mismatches: {diffs[diffs>0].head().to_dict()}"


def test_is_polygon_friendly_rule_applied_correctly(supercluster_df) -> None:
    """is_polygon_friendly = (count_ways + count_relations) / count_all >= 0.5.
    Re-derive the boolean for every row and compare."""
    expected = (
        (supercluster_df["count_ways"] + supercluster_df["count_relations"])
        / supercluster_df["count_all"]
        >= POLYGON_FRIENDLY_THRESHOLD
    )
    assert (supercluster_df["is_polygon_friendly"] == expected).all(), (
        f"is_polygon_friendly mismatches: "
        f"{(supercluster_df['is_polygon_friendly'] != expected).sum()} rows differ"
    )


def test_yes_labeled_base_keys_all_have_at_least_one_cluster(supercluster_df) -> None:
    """Every yes-labeled base key should map to at least one real cluster."""
    zero_rows = supercluster_df[supercluster_df["count_all"] == 0]
    assert len(zero_rows) == 0, (
        f"{len(zero_rows)} yes-labeled base keys have zero clusters: "
        f"{zero_rows['base_key'].tolist()[:10]}"
    )


def test_no_zero_rows_for_yes_labels(supercluster_df, yes_labels) -> None:
    """One row per yes label, all with non-zero cluster count."""
    assert len(supercluster_df) == len(yes_labels)
    assert (supercluster_df["n_clusters"] > 0).all()


def test_headline_numbers(supercluster_df) -> None:
    """Lock the headline numbers reported to the user.

    157 yes-labeled base keys:
      - 110 polygon-friendly, 47 point-heavy
      - 1,061,684,644 polygon-friendly occurrences
      -   105,791,583 point-heavy occurrences
      -     12,504 polygon-friendly tags
      -      2,354 point-heavy tags
      -        670 polygon-friendly real clusters
      -        189 point-heavy real clusters
    """
    n_pf = int(supercluster_df["is_polygon_friendly"].sum())
    n_ph = int((~supercluster_df["is_polygon_friendly"]).sum())
    assert n_pf == 110, f"polygon-friendly: expected 110, got {n_pf}"
    assert n_ph == 47, f"point-heavy: expected 47, got {n_ph}"

    occ_pf = int(supercluster_df.loc[supercluster_df["is_polygon_friendly"], "count_all"].sum())
    occ_ph = int(supercluster_df.loc[~supercluster_df["is_polygon_friendly"], "count_all"].sum())
    assert occ_pf == 1_061_684_644, f"occ PF: expected 1,061,684,644, got {occ_pf:,}"
    assert occ_ph == 105_791_583, f"occ PH: expected 105,791,583, got {occ_ph:,}"

    tags_pf = int(supercluster_df.loc[supercluster_df["is_polygon_friendly"], "n_tags"].sum())
    tags_ph = int(supercluster_df.loc[~supercluster_df["is_polygon_friendly"], "n_tags"].sum())
    assert tags_pf == 12_504, f"tags PF: expected 12,504, got {tags_pf:,}"
    assert tags_ph == 2_354, f"tags PH: expected 2,354, got {tags_ph:,}"

    clusters_pf = int(supercluster_df.loc[supercluster_df["is_polygon_friendly"], "n_clusters"].sum())
    clusters_ph = int(supercluster_df.loc[~supercluster_df["is_polygon_friendly"], "n_clusters"].sum())
    assert clusters_pf == 670, f"clusters PF: expected 670, got {clusters_pf:,}"
    assert clusters_ph == 189, f"clusters PH: expected 189, got {clusters_ph:,}"


def test_percentages_sum_to_100(supercluster_df) -> None:
    """pct_nodes + pct_ways + pct_relations ≈ 100 for every row.

    The percentages are computed as count_nodes / count_all * 100,
    etc. The source DB's count_all and count_nodes + count_ways +
    count_relations can differ by a few occurrences per (key, value)
    due to how taginfo aggregates case variants; the total drift per
    supercluster is small (max 146 occurrences observed in 2026-06).
    We allow up to 0.3 percentage points of drift.
    """
    total_pct = (
        supercluster_df["pct_nodes"]
        + supercluster_df["pct_ways"]
        + supercluster_df["pct_relations"]
    )
    bad = total_pct[(total_pct < 99.7) | (total_pct > 100.3)]
    assert len(bad) == 0, (
        f"{(total_pct < 99.7).sum() + (total_pct > 100.3).sum()} rows "
        f"with bad pct total: {bad.head().to_dict()}"
    )


def test_total_clusters_across_superclusters_matches_unique_real_clusters(
    supercluster_df, real_memberships,
) -> None:
    """Sum of n_clusters across all 157 superclusters = number of
    distinct real clusters whose medoid has a yes-labeled base key.
    """
    actual = int(supercluster_df["n_clusters"].sum())
    expected = real_memberships["cluster_id"].nunique()
    assert actual == expected, f"actual={actual}, expected={expected}"
